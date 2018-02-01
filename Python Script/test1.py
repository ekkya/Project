import xlsxwriter
import openpyxl
import os.path
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from scipy import signal
from xlrd import open_workbook

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook('data_L23_1.xlsx')
worksheet = workbook.add_worksheet()

# Some data we want to write to the worksheet.
row = 0
col = 0

with open('soma_voltage_L23_1.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1

print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L23_2.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L23_2.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1

print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L23_3.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L23_3.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L4_1.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L4_1.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L4_2.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L4_2.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L4_3.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L4_3.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L5_1.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L5_1.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L5_2.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L5_2.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L5_3.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L5_3.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L6_1.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L6_1.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L6_2.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L6_2.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

workbook = xlsxwriter.Workbook('data_L6_3.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0

with open('soma_voltage_L6_3.dat', 'r') as f:
    data = f.readlines()
    #print data
    for line in data:
        words = line.split()
        worksheet.write(row, col, words[0])
        worksheet.write(row, col + 1, words[1])
        row += 1


print "Complete"
workbook.close()

