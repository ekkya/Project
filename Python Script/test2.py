import xlsxwriter
import openpyxl
import os.path
import sys
from scipy import signal
from openpyxl import Workbook
from openpyxl import load_workbook
from xlrd import open_workbook

row = 0
col = 0
list = []
wb = open_workbook('data_L23_1.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L23_1_u.xlsx')
worksheet = workbook.add_worksheet()
i = 0
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        #print s
        list.append(s[0])
        i = i + 1
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L23')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)


print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L23_2.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L23_2_u.xlsx')
worksheet = workbook.add_worksheet()
i = 0
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        i = i + 1
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L23')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)

#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L23_3.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L23_3_u.xlsx')
worksheet = workbook.add_worksheet()
i = 0
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        i = i + 1
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L23')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L4_1.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L4_1_u.xlsx')
worksheet = workbook.add_worksheet()
i = 0
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        i = i + 1
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L4')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L4_2.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L4_2_u.xlsx')
worksheet = workbook.add_worksheet()
i = 0
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        i = i + 1
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L4')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L4_3.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L4_3_u.xlsx')
worksheet = workbook.add_worksheet()
i = 0
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        i = i + 1
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L4')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L5_1.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L5_1_u.xlsx')
worksheet = workbook.add_worksheet()
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L5')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L5_2.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L5_2_u.xlsx')
worksheet = workbook.add_worksheet()
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L5')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L5_3.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L5_3_u.xlsx')
worksheet = workbook.add_worksheet()
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L5')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L6_1.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L6_1_u.xlsx')
worksheet = workbook.add_worksheet()
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L6')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L6_2.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L6_2_u.xlsx')
worksheet = workbook.add_worksheet()
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L6')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()

row = 0
col = 0
list = []
wb = open_workbook('data_L6_3.xlsx')
sheet = wb.sheet_by_index(0)
workbook = xlsxwriter.Workbook('data_L6_3_u.xlsx')
worksheet = workbook.add_worksheet()
for row_num in range(sheet.nrows):
    row_value = sheet.row_values(row_num)
    if row_value[1] > "0":
        #print row_value
        s = [str(item) for item in row_value]
        print s
        list.append(s[0])
        worksheet.write(row, col, s[0])
        worksheet.write(row, col + 1, s[1])
        worksheet.write(row, col + 2, 'L6')
        worksheet.write(row, col + 3, 1 / (float(s[0]) * 0.001))
        row += 1
    #row = 1
    worksheet.write(row, col + 4, i)
#print len(list)
results = map(float, list)
#print len(results)
r,fx = signal.periodogram(results)
i = sum(fx)/len(fx)
row = 0
col = 0
worksheet.write(row, col + 7, i)

print "Complete"
workbook.close()